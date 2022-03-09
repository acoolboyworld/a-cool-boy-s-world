from openpyxl import Workbook,load_workbook

# wb = Workbook()  #创建了一个excel
# sheet = wb.active
# sheet.title = '呵呵'
# sheet['A1'] = '爸爸'
#
# wb.save('python处理excel.xlsx')
wb = load_workbook('python处理excel.xlsx')
sheet = wb.get_sheet_by_name("呵呵")


print(sheet['A1'].value)
